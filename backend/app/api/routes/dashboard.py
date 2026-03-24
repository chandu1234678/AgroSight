from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from app.api.deps import get_current_user
from app.database import get_db
from app.models.user import User
from app.models.scan import Scan
from datetime import datetime
import io
import csv

router = APIRouter()


# ── helpers ──────────────────────────────────────────────────────────────────

def _scan_rows(scans):
    rows = []
    for s in scans:
        rows.append({
            "id": s.id,
            "disease": s.disease.replace("_", " ").title() if s.disease else "Unknown",
            "confidence": f"{round((s.confidence or 0) * 100, 1)}%",
            "severity": (s.severity or "N/A").capitalize(),
            "date": s.created_at.strftime("%Y-%m-%d %H:%M") if s.created_at else "N/A",
            "recommendation": (s.recommendation or "").replace("\n", " "),
        })
    return rows


def _build_csv(rows, user_name: str) -> io.BytesIO:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["AgroSight — Diagnostic Report"])
    w.writerow([f"User: {user_name}", f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"])
    w.writerow([])
    w.writerow(["Scan ID", "Disease / Condition", "Confidence", "Severity", "Date", "Recommendation"])
    for r in rows:
        w.writerow([r["id"], r["disease"], r["confidence"], r["severity"], r["date"], r["recommendation"]])
    return io.BytesIO(buf.getvalue().encode("utf-8"))


def _build_excel(rows, stats: dict, user_name: str) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Diagnostic Report"

    # ── colour palette ──
    GREEN_DARK  = "1A2E1A"
    GREEN_MID   = "2D5A27"
    GREEN_LIGHT = "4CAF50"
    ACCENT      = "A8D5A2"
    BG_ROW_ODD  = "F4FAF4"
    BG_ROW_EVEN = "FFFFFF"
    RED_SOFT    = "FFEBEE"
    RED_TEXT    = "C62828"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title block ──
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "🌿  AgroSight — Diagnostic Intelligence Report"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=GREEN_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = f"User: {user_name}   |   Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
    sub.font = Font(name="Calibri", size=10, color="FFFFFF", italic=True)
    sub.fill = PatternFill("solid", fgColor=GREEN_MID)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # ── Summary stats ──
    ws.row_dimensions[3].height = 8
    stat_labels = ["Total Scans", "Diseases Detected", "Healthy Scans", "Healthy %", "Avg Accuracy"]
    stat_values = [
        stats.get("total_scans", 0),
        stats.get("diseases_detected", 0),
        stats.get("healthy_scans", 0),
        f"{stats.get('healthy_pct', 0)}%",
        f"{stats.get('accuracy', 0)}%",
    ]
    for col, (lbl, val) in enumerate(zip(stat_labels, stat_values), start=1):
        lc = ws.cell(row=4, column=col, value=lbl)
        lc.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        lc.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
        lc.alignment = Alignment(horizontal="center")
        lc.border = border

        vc = ws.cell(row=5, column=col, value=val)
        vc.font = Font(name="Calibri", bold=True, size=13)
        vc.alignment = Alignment(horizontal="center")
        vc.border = border
    ws.row_dimensions[5].height = 24

    # ── Table header ──
    ws.row_dimensions[6].height = 8
    headers = ["Scan ID", "Disease / Condition", "Confidence", "Severity", "Date", "Recommendation"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=7, column=col, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=GREEN_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[7].height = 22

    # ── Data rows ──
    for i, r in enumerate(rows):
        row_num = 8 + i
        bg = BG_ROW_ODD if i % 2 == 0 else BG_ROW_EVEN
        is_disease = "healthy" not in r["disease"].lower()
        values = [r["id"], r["disease"], r["confidence"], r["severity"], r["date"], r["recommendation"]]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.font = Font(name="Calibri", size=10,
                          color=RED_TEXT if (is_disease and col == 2) else "000000")
            c.fill = PatternFill("solid", fgColor=RED_SOFT if is_disease else bg)
            c.alignment = Alignment(vertical="center", wrap_text=(col == 6))
            c.border = border
        ws.row_dimensions[row_num].height = 18

    # ── Column widths ──
    col_widths = [10, 28, 14, 14, 20, 60]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_pdf(rows, stats: dict, user_name: str) -> io.BytesIO:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=15*mm, bottomMargin=15*mm,
    )

    GREEN  = colors.HexColor("#1B5E20")
    GREEN2 = colors.HexColor("#4CAF50")
    LIGHT  = colors.HexColor("#E8F5E9")
    RED    = colors.HexColor("#C62828")
    RED_BG = colors.HexColor("#FFEBEE")
    GREY   = colors.HexColor("#F5F5F5")
    WHITE  = colors.white

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", fontSize=18, textColor=WHITE,
                                  alignment=TA_CENTER, fontName="Helvetica-Bold", spaceAfter=2)
    sub_style   = ParagraphStyle("sub",   fontSize=9,  textColor=colors.HexColor("#A5D6A7"),
                                  alignment=TA_CENTER, fontName="Helvetica", spaceAfter=0)
    cell_style  = ParagraphStyle("cell",  fontSize=8,  textColor=colors.black,
                                  fontName="Helvetica", leading=11)
    cell_red    = ParagraphStyle("cellr", fontSize=8,  textColor=RED,
                                  fontName="Helvetica-Bold", leading=11)

    story = []

    # Title
    title_tbl = Table([[
        Paragraph("🌿  AgroSight — Diagnostic Intelligence Report", title_style),
    ]], colWidths=[doc.width])
    title_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), GREEN),
        ("TOPPADDING",    (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("ROUNDEDCORNERS", [4]),
    ]))
    story.append(title_tbl)

    sub_tbl = Table([[
        Paragraph(f"User: {user_name}   |   Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC", sub_style),
    ]], colWidths=[doc.width])
    sub_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#2E7D32")),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(sub_tbl)
    story.append(Spacer(1, 6*mm))

    # Stats summary
    stat_headers = ["Total Scans", "Diseases Detected", "Healthy Scans", "Healthy %", "Avg Accuracy"]
    stat_vals    = [
        str(stats.get("total_scans", 0)),
        str(stats.get("diseases_detected", 0)),
        str(stats.get("healthy_scans", 0)),
        f"{stats.get('healthy_pct', 0)}%",
        f"{stats.get('accuracy', 0)}%",
    ]
    stat_col = doc.width / 5
    stat_tbl = Table(
        [[Paragraph(h, ParagraphStyle("sh", fontSize=8, textColor=WHITE, fontName="Helvetica-Bold", alignment=TA_CENTER)) for h in stat_headers],
         [Paragraph(v, ParagraphStyle("sv", fontSize=14, textColor=WHITE, fontName="Helvetica-Bold", alignment=TA_CENTER)) for v in stat_vals]],
        colWidths=[stat_col]*5,
    )
    stat_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), GREEN2),
        ("BACKGROUND", (0,1), (-1,1), colors.HexColor("#388E3C")),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#C8E6C9")),
    ]))
    story.append(stat_tbl)
    story.append(Spacer(1, 6*mm))

    # Table
    col_widths_pdf = [18*mm, 55*mm, 22*mm, 22*mm, 35*mm, None]
    col_widths_pdf[-1] = doc.width - sum(w for w in col_widths_pdf[:-1])

    header_row = [
        Paragraph(h, ParagraphStyle("th", fontSize=9, textColor=WHITE,
                                     fontName="Helvetica-Bold", alignment=TA_CENTER))
        for h in ["ID", "Disease / Condition", "Confidence", "Severity", "Date", "Recommendation"]
    ]
    data = [header_row]
    for r in rows:
        is_disease = "healthy" not in r["disease"].lower()
        data.append([
            Paragraph(str(r["id"]), cell_style),
            Paragraph(r["disease"], cell_red if is_disease else cell_style),
            Paragraph(r["confidence"], cell_style),
            Paragraph(r["severity"], cell_style),
            Paragraph(r["date"], cell_style),
            Paragraph(r["recommendation"][:200] + ("…" if len(r["recommendation"]) > 200 else ""), cell_style),
        ])

    tbl = Table(data, colWidths=col_widths_pdf, repeatRows=1)
    row_styles = [
        ("BACKGROUND", (0,0), (-1,0), GREEN),
        ("TEXTCOLOR",  (0,0), (-1,0), WHITE),
        ("GRID",       (0,0), (-1,-1), 0.4, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("VALIGN",     (0,0), (-1,-1), "TOP"),
    ]
    for i in range(1, len(data)):
        is_d = "healthy" not in rows[i-1]["disease"].lower()
        bg = RED_BG if is_d else (GREY if i % 2 == 0 else WHITE)
        row_styles.append(("BACKGROUND", (0,i), (-1,i), bg))
    tbl.setStyle(TableStyle(row_styles))
    story.append(tbl)

    doc.build(story)
    buf.seek(0)
    return buf


# ── routes ────────────────────────────────────────────────────────────────────

@router.get("/stats")
async def get_dashboard_stats(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    total_result = await db.execute(
        select(func.count(Scan.id)).where(Scan.user_id == current_user.id)
    )
    total_scans = total_result.scalar() or 0

    all_result = await db.execute(
        select(Scan).where(Scan.user_id == current_user.id)
    )
    all_scans = all_result.scalars().all()

    diseases_detected = sum(1 for s in all_scans if s.disease and "healthy" not in s.disease.lower())
    healthy_count     = sum(1 for s in all_scans if s.disease and "healthy" in s.disease.lower())
    accuracy = round(sum(s.confidence or 0 for s in all_scans) / len(all_scans) * 100, 1) if all_scans else 0.0
    healthy_pct = round(healthy_count / total_scans * 100) if total_scans > 0 else 0

    most_common_result = await db.execute(
        select(Scan.disease, func.count(Scan.disease).label("count"))
        .where(Scan.user_id == current_user.id)
        .group_by(Scan.disease)
        .order_by(func.count(Scan.disease).desc())
        .limit(1)
    )
    most_common = most_common_result.first()

    return {
        "total_scans": total_scans,
        "diseases_detected": diseases_detected,
        "healthy_scans": healthy_count,
        "healthy_pct": healthy_pct,
        "accuracy": accuracy,
        "most_common_disease": most_common[0] if most_common else None,
    }


@router.get("/report/download")
async def download_report(
    format: str = Query("csv", regex="^(csv|excel|pdf)$"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Download scan report as CSV, Excel, or PDF."""
    result = await db.execute(
        select(Scan).where(Scan.user_id == current_user.id).order_by(Scan.created_at.desc())
    )
    scans = result.scalars().all()

    # Also fetch stats for summary section
    all_scans = scans
    total_scans = len(all_scans)
    diseases_detected = sum(1 for s in all_scans if s.disease and "healthy" not in s.disease.lower())
    healthy_count     = sum(1 for s in all_scans if s.disease and "healthy" in s.disease.lower())
    accuracy = round(sum(s.confidence or 0 for s in all_scans) / len(all_scans) * 100, 1) if all_scans else 0.0
    healthy_pct = round(healthy_count / total_scans * 100) if total_scans > 0 else 0

    stats = {
        "total_scans": total_scans,
        "diseases_detected": diseases_detected,
        "healthy_scans": healthy_count,
        "healthy_pct": healthy_pct,
        "accuracy": accuracy,
    }

    rows = _scan_rows(scans)
    date_str = datetime.utcnow().strftime("%Y%m%d")
    safe_name = current_user.name.replace(" ", "_")

    if format == "excel":
        buf = _build_excel(rows, stats, current_user.name)
        filename = f"agrosight_report_{safe_name}_{date_str}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif format == "pdf":
        buf = _build_pdf(rows, stats, current_user.name)
        filename = f"agrosight_report_{safe_name}_{date_str}.pdf"
        media_type = "application/pdf"
    else:
        buf = _build_csv(rows, current_user.name)
        filename = f"agrosight_report_{safe_name}_{date_str}.csv"
        media_type = "text/csv"

    return StreamingResponse(
        buf,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
